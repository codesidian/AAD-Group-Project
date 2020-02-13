from background_task import background
from django.conf import settings
from api.models import Notification, Sale, SaleItem, Item, Customer, Return, Staff, StockCheck, StockCheckItem, Report
from openpyxl import Workbook
import uuid
from django.utils import timezone

REPORT_DIR = "staff/reports/"

@background(schedule=60)
def generateSalesReport(userid, fromDate, toDate):
    periodOfSales = Sale.objects.filter(datetime__range=[fromDate, toDate])
    headers = ['ID', 'Date', 'Customer Name', 'Quantity', 'Price']
    wb = Workbook()
    ws1 = wb.active
    # Sheet 1 sales report overview
    ws1.title = "Sales Report Overview"
    ws1.append(headers)
    for sale in periodOfSales:
        saleItems = SaleItem.objects.filter(sale_id=sale.id)
        currentPrice = 0
        currentQuantity = 0
        for items in saleItems:
            currentPrice = currentPrice + (items.sale_price * items.quantity)
            currentQuantity = currentQuantity + items.quantity
        customer = Customer.objects.get(user_id=sale.customer_id)
        customerName = customer.full_name
        currentRow = [
            sale.id, sale.datetime, customerName, currentQuantity, currentPrice
        ]
        ws1.append(currentRow)
    # Sheet 2 sales report details
    headers = [
        'Sale ID', 'Item Code', 'Item Name', 'Quantity', 'Price', 'Returned',
        'Date', 'Department', 'Customer Name'
    ]
    ws2 = wb.create_sheet(title="Weekly Sales Item Report")
    ws2.title = "Sales Report Details"
    ws2.append(headers)
    for sale in periodOfSales:
        saleItems = SaleItem.objects.filter(sale_id=sale.id)
        currentPrice = 0
        currentQuantity = 0
        customer = Customer.objects.get(user_id=sale.customer_id)
        for items in saleItems:
            itemDetails = Item.objects.get(id=items.item_id)
            currentRow = [
                sale.id, itemDetails.code, itemDetails.name, items.quantity,
                items.sale_price, items.returned_quantity, sale.datetime,
                customer.dept_id, customer.full_name
            ]
            ws2.append(currentRow)
    filename = "Sales_Report"+str(timezone.now().strftime("%Y%m%d-%H%M%S"))+".xlsx"
    wb.save(REPORT_DIR+filename)
    
    reportMessage="Sales Report Ready"
    report = Report(user_id=userid,
                filename=filename,
                created_date=timezone.now(),
                report_type="ST")
    report.save()    
    notify = Notification(user_id=userid,
                          text=reportMessage,
                          notification_type="RE",
                          link="/staff/reports/?id="+str(report.id),
                          seen=False)
    notify.save()
    

@background(schedule=60)
def generateReturnsReport(userid, fromDate, toDate):
    periodOfReturns = Return.objects.filter(datetime__range=[fromDate, toDate])
    headers = [
        'Return ID', 'Sale ID', 'Item ID', 'Item Name', 'Customer Name',
        'Staff Name', 'Quantity', 'Reason', 'Date'
    ]
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Return Report"
    ws1.append(headers)
    for ret in periodOfReturns:
        saleItem = SaleItem.objects.get(id=ret.sale_item_id)
        item = Item.objects.get(id=saleItem.item_id)
        sale = Sale.objects.get(id=saleItem.sale_id)
        staff = Staff.objects.get(user_id=ret.staff_id)
        customer = Customer.objects.get(user_id=sale.customer_id)
        currentRow = [
            ret.id, sale.id, item.id, item.name, customer.full_name,
            staff.full_name, ret.quantity, ret.reason, ret.datetime
        ]
        ws1.append(currentRow)
    filename = "Return_Report"+str(timezone.now().strftime("%Y%m%d-%H%M%S"))+".xlsx"
    wb.save(REPORT_DIR+filename)
    reportMessage="Return Report Ready"
    
    report = Report(user_id=userid,
                    filename=filename,
                    created_date=timezone.now(),
                    report_type="ST")
    report.save()
    notify = Notification(user_id=userid,
                          text=reportMessage,
                          notification_type="RE",
                          link="/staff/reports/?id="+str(report.id),
                          seen=False)
    notify.save()


# We can take a bool for if to include a second sheet for stockchecks.
@background(schedule=60)
def generateStockReport(userid, includeChecks, fromDate="", toDate=""):
    items = Item.objects.all()
    headers = [
        'Item ID', 'Item Code', 'Item Name', 'Item Price', 'Quantity',
        'Warning Quantity', 'Is Chemical', 'Pack Size', 'For Sale'
    ]
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Stock Report"
    ws1.append(headers)
    for item in items:
        currentRow = [
            item.id, item.code, item.name, (item.price / 100), item.quantity,
            item.warning_quantity, item.is_chemical, item.pack_size,
            item.for_sale
        ]
        ws1.append(currentRow)
    # Optional Stock Checks
    if includeChecks:
        # Sheet 2 stocks checks
        headers = [
            'Date', 'Staff Member', 'Item Code', 'Item Name',
            'Observed Quantity', 'Expected Quantity', 'Warning Quantity'
        ]
        ws2 = wb.create_sheet(title="Stock Check Report")
        ws2.title = "Stock Check Report"
        ws2.append(headers)
        for check in StockCheck.objects.filter(
                datetime__range=[fromDate, toDate]):
            checkStaff = Staff.objects.get(user_id=check.staff_id)
            checkedItem = StockCheckItem.objects.get(id=check.id)
            itemDetails = Item.objects.get(id=checkedItem.item_id)
            currentRow = [
                check.datetime, checkStaff.full_name, itemDetails.code,
                itemDetails.name, checkedItem.observed_quantity,
                checkedItem.expected_quantity, itemDetails.warning_quantity
            ]
            ws2.append(currentRow)
    filename = "Stock_Report"+str(timezone.now().strftime("%Y%m%d-%H%M%S"))+".xlsx"
    wb.save(REPORT_DIR+filename)
    reportMessage="Stock Report Ready"
    
    report = Report(user_id=userid,
                    filename=filename,
                    created_date=timezone.now(),
                    report_type="ST")
    report.save()
    notify = Notification(user_id=userid,
                          text=reportMessage,
                          notification_type="RE",
                          link="/staff/reports/?id="+str(report.id),
                          seen=False)
    notify.save()


def resetPasswordEmail(userid):
    # ask for email
    # new pass
    return NotImplementedError